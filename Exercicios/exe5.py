from openpyxl import Workbook 

palavras = ['Adaylson', 'Progamação', 'Vida', 'Teste', 'dois', 'um', 'maio', 'novembro', 'gaita', 'ultimo']

def stutter(palavra):
    duas_primeiras = palavra[0:2]
    return duas_primeiras + '... ' + duas_primeiras + '... ' + palavra + '?'


wb = Workbook()
ws = wb.active
ws.title = 'Gaguejar'

ws['A1'] = 'Guaguejar palavras'
ws['A2'] = 'Palavra'
ws['B2'] = 'Gaguejada'

for x in range(3,13):
    for y in range(1,3):
        if(y == 1):
            ws.cell(row=x, column=y, value=palavras[x-3])
        else:
            ws.cell(row=x, column=y, value=stutter(palavras[x-3]))

wb.save('exe.xlsx')