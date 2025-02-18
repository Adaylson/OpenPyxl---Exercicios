from openpyxl import Workbook
import random

wb = Workbook()
ws = wb.active
ws.title = 'Minutos para segundos'

def conversor(minutos):
    return minutos*60

ws['A1'] = 'Conversor'
ws['A2'] = 'Minutos'
ws['B2'] = 'Segundos'

for x in range(3,13):
    for y in range(1,3):
        if(y==1):
            ws.cell(row=x, column=y, value=random.randint(1, 100))
        else:
            ws.cell(row=x, column=y, value=(conversor(ws.cell(row=x,column=y-1).value)))

wb.save('exe.xlsx')