from openpyxl import Workbook
import random

wb = Workbook()
ws = wb.active
ws.title = "Operador modulo"

ws['A1'] = "Retornando o resultado"
ws['A2'] = "Dividendo"
ws['B2'] = "Divisor"
ws['C2'] = "Consiente"

def remainder(dividendo, divisor):
    return dividendo % divisor

for x in range(3,13):
    for y in range(1,4):
        if(y == 1):
            ws.cell(row=x, column=y, value=random.randint(1,50))
        elif(y == 2):
            ws.cell(row=x, column=y, value=random.randint(1,40))
        else:
            ws.cell(row=x, column=y, value=remainder(ws.cell(row=x, column=y-2).value, ws.cell(row=x, column=y-1).value))

wb.save("exe4.xlsx")