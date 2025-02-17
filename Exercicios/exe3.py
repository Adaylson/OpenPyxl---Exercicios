from openpyxl import Workbook
import random
import datetime

wb = Workbook()
ws = wb.active
ws.title = "Dias de vida"

ws['A1'] = 'Dias de vida'
ws['A2'] = 'Ano'
ws['B2'] = 'Mes'
ws['C2'] = 'Dia'
ws['D2'] = 'Tempo de vida expresso em dias'

data = datetime.date.today()

diaatual = data.day
mesatual = data.month
anoatual = data.year


for x in range(3,13):
    for y in range(1,5):
        if(y==1):
            ws.cell(row=x,column=y,value=random.randint(1900, 2025))
        elif(y==2):
            ws.cell(row=x,column=y,value=random.randint(1,12))
        elif(y==3):
            ws.cell(row=x,column=y,value=random.randint(1,30))
        else:
            ws.cell(row=x,column=y,value=( ((anoatual - ws.cell(row=x,column=y-3).value) * 365) +  ((mesatual - ws.cell(row=x,column=y-2).value) * 30) + ((diaatual - ws.cell(row=x,column=y-1).value) * 30) ))

wb.save('exe3.xlsx')