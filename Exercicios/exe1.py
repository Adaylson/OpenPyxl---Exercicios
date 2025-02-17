#Gerar um numero aleatorio para 10 celululas e na frente colocar, o seu antecessor, e seu sucessor

from openpyxl import Workbook
import random

wb = Workbook()

ws = wb.active

ws['A1'] = 'Função do antecessor e sucessor'
ws['A2'] = 'Numero'
ws['B2'] = 'Antecessor'
ws['C2'] = 'Sucessor'
for x in range(3, 13):
    for y in range(1,4):

        if(y==1):
            ws.cell(row=x, column=y, value=random.randint(1,100))
        elif(y==2):
            ws.cell(row=x, column=y, value=ws.cell(row=x, column=y-1,).value - 1)
        else:
            ws.cell(row=x, column=y, value=ws.cell(row=x, column=y-1,).value + 1)

wb.save('exe1.xlsx')