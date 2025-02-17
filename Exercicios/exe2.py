#Fazer uma tabela 10x3, gerando base altura e a area de um retangulo

from openpyxl import Workbook
import random

wb = Workbook()
ws = wb.active

ws.title = 'Area Retangulo'
ws['A1']='Area retangulo'
ws['A2']='Base'
ws['B2']='Altura'
ws['C2']='Area'


for x in range(3, 13):
    for y in range(1,4):
        if(y!=3):
            ws.cell(row=x, column=y, value=random.randint(1,100))
        else:
            ws.cell(row=x, column=y, value= ws.cell(row=x, column=y-2).value*ws.cell(row=x, column=y-1).value )

wb.save('exe2.xlsx')
