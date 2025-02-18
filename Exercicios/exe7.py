# Não vai funcionar por conta do erro de solução, no caso, posso dividir o array em mais celulas ou deixar o valor separado

from openpyxl import Workbook
import random

wb = Workbook()
ws = wb.active
ws.title = 'arestas'

# Gerar array 
def gerador():
    array = [[]]
    for x in range(0,5):
        temp = []
        for y in range(0,5):
            temp.insert(y,random.randint(0,1))
        array.insert(x,temp)
    return array

def Verificar(array, x, y):
    if (array[x][y])==1:
        return(True)
    else:
        return(False)
    
for x in range(3,13):
    for y in range(1,3):
        if(y==1):
            ws.cell(row=x, column=y, value=gerador())
        else:
            ws.cell(row=x, column=y, value=Verificar(ws.cell(row=x, column=y-1).value))

wb.save('exe.xlsx')
